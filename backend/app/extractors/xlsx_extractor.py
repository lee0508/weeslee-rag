"""
XLSX Extractor
"""
import os
from typing import Dict, Any, List
from pathlib import Path
from openpyxl import load_workbook

from app.extractors.base import BaseExtractor, ExtractionResult


class XlsxExtractor(BaseExtractor):
    """Extracts text from XLSX files"""

    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".xls"]

    async def extract(self, file_path: str) -> Dict[str, Any]:
        """
        Extract text from XLSX file

        Args:
            file_path: Path to XLSX file

        Returns:
            Extraction result dictionary
        """
        if not os.path.exists(file_path):
            return ExtractionResult(
                success=False,
                error=f"File not found: {file_path}"
            ).to_dict()

        try:
            wb = load_workbook(file_path, data_only=True)
            content_parts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_content = [f"--- Sheet: {sheet_name} ---"]

                for row in sheet.iter_rows():
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            row_values.append(str(cell.value))
                        else:
                            row_values.append("")

                    # Only add non-empty rows
                    if any(v.strip() for v in row_values):
                        sheet_content.append(" | ".join(row_values))

                content_parts.append("\n".join(sheet_content))

            content = "\n\n".join(content_parts)

            metadata = {
                "source": file_path,
                "filename": Path(file_path).name,
                "sheets": wb.sheetnames,
                "sheet_count": len(wb.sheetnames)
            }

            wb.close()

            return ExtractionResult(
                success=True,
                content=content,
                metadata=metadata,
                method="openpyxl"
            ).to_dict()

        except Exception as e:
            return ExtractionResult(
                success=False,
                error=str(e),
                method="openpyxl"
            ).to_dict()


# Singleton instance
xlsx_extractor = XlsxExtractor()
